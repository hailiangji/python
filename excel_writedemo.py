#from openpyxl import load_workbook
from openpyxl import Workbook


if __name__ == '__main__':
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = 'Devansh Sharma'
    sheet['A2'] = 'hello world'
    sheet.cell(row=2, column=2).value = "AAA"
    wb.save(r'demo.xlsx')
    print("运行结束！")