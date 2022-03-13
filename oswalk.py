# -*- coding: utf-8 -*-
import os
import time
import pandas as pd
import pdfinfo
import csv

file_dir = r"C:\Users\HailiangJi\Desktop\ADI-混合信号电子系统PDF"


def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        print("file root =",root)  # 当前目录路径
        print("file dir =",dirs)  # 当前路径下所有子目录
        print("files =",files)  # 当前路径下所有非目录子文件
    # print("type root",type(root))
    # print("type dir",type(dirs))
    # print("type files",type(files))
    L = [] #文件路径+文件名
    name = [] #文件名
    for file in files :
        #print(file)
        name.append(file)
        file_extension = os.path.splitext(file)[-1] #提取文件后缀
        #print(file_extension, type(file_extension ))
        if file_extension == ".pdf" :
            L.append(os.path.join(root, file))  #拼接文档路径和文件名，并存入列表
    #print(L)
    return L, name  #以list形式返回

import sys
import time
import warnings
from openpyxl import Workbook

if __name__ == "__main__" :
    wb = Workbook() #创建一个工作表
    sheet = wb.active #
    filedir,names=file_name(file_dir)
    csv_file = r'C:\Users\HailiangJi\python\aa.csv'
    ff = 1
    for flist  in filedir  :

        dictpdf=pdfinfo.extract_content(flist)  #输入文件，返回pdf文档信息，返回的格式字典
        i=2

        keylist = list(dictpdf.keys())
        valuelist=list(dictpdf.values())
        sheet.cell(row=ff+1, column=1).value = names[ff-1]
        print('names[ff-1]',names[ff-1])
        sheet.cell(row=1, column=1).value = "pdf"
        if ff == 1:
            ii = 2
            for key in keylist:
                print('key',key)
                sheet.cell(row=1, column=ii).value = key
                ii += 1
        for value in valuelist:
            print("value",value)
            sheet.cell(row=ff+1, column=i).value = value
            i += 1
        print('i',i)
        print("ii", ii)
        ff += 1
        wb.save(r'demo.xlsx')


    r'''
        with open(csv_file,'w+',newline="") as outcsv :
            fwrite = csv.writer(outcsv)
            for rowlist in keylist :
                print('rowlist',rowlist)
                print('type rowlist',type(rowlist))
                print('string rowlist',rowlist)
                fwrite.writerow([str(rowlist)])
                fwrite.writerow((","))
        outcsv.close()
        #break
'''


r'''
        newdict = {}
        for key in dictpdf :
            print("key",key)

            print("type key ",type(str(key)))
            value=[dictpdf['{}'.format(key)]]#将value 转为list
            print("value",value,type(value))


            newdict={str(key): [value]}
            print('new dict',newdict)
            datafram = pd.DataFrame(newdict)#,index=1,columns=1
            print('datafram=====',datafram)
            datafram.to_csv(csv_file, sep=",", mode="a+", index=False,header=False)  # mode = a+ w+
            break
        aaa=input("hhh")
'''
r'''
    print("dictpdf ==== ",dictpdf)
    print("type dictpdf",type(dictpdf))
    #newdict={"item": [i for i in range(len(filedir))],"pdf":filedir,"name":name}.update(dictpdf)

    testdic={'/ADI#20Prelim': 'Mini Tutorial', '/ADI#20Pub#20Year': '2012', '/ADI#20Pubcode': 'MT10902-0-8/12(0)',
     '/ADI#20Rev': '0', '/ADI#20Template': 'UG 1.1', '/ADI#20Title': 'MT-229', '/Author': 'Analog Devices, Inc.',
     '/CreationDate': "D:20120822075836-04'00'", '/Creator': 'Acrobat PDFMaker 10.1 for Word',
     '/Keywords': b'A/D converters\r\nanalog-to-digital converters\r\nADC circuit',
     '/ModDate': "D:20120822082133-04'00'", '/Producer': 'Adobe PDF Library 10.0',
     '/SourceModified': 'D:20120822115822',
     '/Subject': b'Quantization Noise: An Expanded Derivation of the Equation, \r\nSNR = 6.02 N + 1.76 dB\r\n',
     '/Title': 'MT-229'}

    datafram = pd.DataFrame(testdic)
    datafram.to_csv(csv_file,sep=",",mode="w+",index=False)  #mode = a+续写 w+

    # print("TYPE A", type(A))
    # delete file

    # if os.path.exists(csv_file):
    #     os.remove(csv_file)
    #     print("file has been removed")
    # else:
    #     print("no such file %s" % csv_file)

    #df = pd.read_csv(file_dir,sep=" ",engine='python',encoding='ANSI')#gb2312
    #print("df",df)

    #test ok
    filedir,name=file_name(file_dir)
    csv_file = r'C:\Users\HailiangJi\python\aa.csv'
    for flist in filedir :
        dictpdf=pdfinfo.extract_content(flist)
    datafram=pd.DataFrame({"item": [i for i in range(len(filedir))],"pdf":filedir,"name":name})
    datafram.to_csv(csv_file,sep=",",mode="a+",index=False)  #mode = a+ w+
    '''